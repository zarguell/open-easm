"""Excel (.xlsx) report generation for OpenEASM."""

from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Color palette — based on our UI design tokens (dark terminal theme)
BG = "1a1a1a"
CARD = "222222"
CARD_ALT = "2a2a2a"
INK = "f2f2f2"
MUTED = "8b949e"
PRIMARY = "00d992"
PRIMARY_DEEP = "10b981"
RED = "ef4444"
BORDER_COLOR = "3d3a39"

SEVERITY_FILL = {
    "critical": "991b1b",
    "high": "ef4444",
    "medium": "f59e0b",
    "low": "8b949e",
    "info": "3d3a39",
}

SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}

_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


def generate_excel_report(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"

    _sheet_summary(ws, data)
    _sheet_action_plan(wb, data)
    _sheet_findings(wb, data)
    _sheet_exposure(wb, data)
    _sheet_certificates(wb, data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_summary(ws, data: dict):
    risk = data.get("executive_risk", {})
    sla = data.get("sla_summary", {})
    findings = data.get("findings", [])

    ws.merge_cells("A1:J2")
    ws["A1"] = "OPENEASM — EXTERNAL ATTACK SURFACE REPORT"
    ws["A1"].font = Font(bold=True, size=18, color=INK)
    ws["A1"].fill = PatternFill("solid", fgColor=PRIMARY_DEEP)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    kpis = [
        ("Target", data.get("target_id", "N/A")),
        ("Executive Score", f"{risk.get('overall_score', 'N/A')} / 100"),
        ("Technical Score", f"{risk.get('technical_score', 'N/A')} / 1000"),
        ("Posture", risk.get("posture", "N/A")),
        ("Risk Level", risk.get("risk_level", "N/A")),
        ("Profile", risk.get("profile", "N/A")),
        ("IPs", data.get("ip_count", 0)),
        ("Hostnames", data.get("hostname_count", 0)),
        ("Findings", len(findings)),
        ("Overdue SLA", sla.get("overdue_count", 0)),
    ]

    row = 4
    for i, (label, value) in enumerate(kpis):
        col = (i % 5) * 2 + 1
        if i == 5:
            row = 7
        cell = ws.cell(row, col, label)
        cell.font = Font(bold=True, color=PRIMARY, size=9)
        cell.fill = PatternFill("solid", fgColor=CARD_ALT)
        cell = ws.cell(row + 1, col, str(value))
        cell.font = Font(bold=True, color=INK, size=13)
        cell.fill = PatternFill("solid", fgColor=CARD)

    # Board summary
    ws.merge_cells("A10:J12")
    ws["A10"] = risk.get("board_summary", "No summary available.")
    ws["A10"].font = Font(size=12, color=INK)
    ws["A10"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A10"].fill = PatternFill("solid", fgColor=CARD)

    # Severity distribution
    by_sev = risk.get("by_severity", {})
    start = 14
    ws.cell(start, 1, "Findings by Severity")
    ws.cell(start, 1).font = Font(bold=True, color=PRIMARY, size=14)
    ws.cell(start + 1, 1, "Severity")
    ws.cell(start + 1, 2, "Count")
    for idx, key in enumerate(["critical", "high", "medium", "low", "info"], start + 2):
        ws.cell(idx, 1, key)
        ws.cell(idx, 2, by_sev.get(key, 0))
        ws.cell(idx, 1).fill = PatternFill("solid", fgColor=SEVERITY_FILL.get(key, MUTED))
        ws.cell(idx, 1).font = Font(bold=True, color="FFFFFF")

    chart = BarChart()
    chart.title = "Findings by Severity"
    data_ref = Reference(ws, min_col=2, min_row=start + 1, max_row=start + 6)
    cats = Reference(ws, min_col=1, min_row=start + 2, max_row=start + 6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, "D14")

    _apply_borders(ws)


def _sheet_action_plan(wb, data: dict):
    ws = wb.create_sheet("Action Plan")
    findings = sorted(
        data.get("findings", []),
        key=lambda f: SEVERITY_ORDER.get(f.get("risk", "info"), 9),
    )

    headers = ["Priority", "Severity", "Rule", "Headline", "SLA (days)", "Status"]
    _write_header(ws, "PRIORITIZED ACTION PLAN", headers)

    sla_policy = {"critical": 3, "high": 14, "medium": 30, "low": 60, "info": None}
    for i, f in enumerate(findings[:50], start=4):
        sev = f.get("risk", "info")
        priority = {"critical": "P1", "high": "P2", "medium": "P3", "low": "P4", "info": "Info"}.get(sev, "Info")
        sla_days = sla_policy.get(sev)
        row = [
            priority, sev, f.get("rule_id", ""), f.get("headline", ""),
            str(sla_days) if sla_days else "-", f.get("status", "open"),
        ]
        for j, val in enumerate(row, 1):
            cell = ws.cell(i, j, val)
            cell.alignment = Alignment(wrap_text=True)
            cell.border = _border

    _apply_borders(ws)


def _sheet_findings(wb, data: dict):
    ws = wb.create_sheet("Findings")
    findings = sorted(
        data.get("findings", []),
        key=lambda f: SEVERITY_ORDER.get(f.get("risk", "info"), 9),
    )

    headers = ["Severity", "Rule", "Headline", "Description", "Status", "Confidence", "First Seen"]
    _write_header(ws, "ALL FINDINGS", headers)

    for i, f in enumerate(findings[:200], start=4):
        evidence = f.get("evidence", {}) or {}
        loc = evidence.get("location", {}) or {}
        row = [
            f.get("risk", ""), f.get("rule_id", ""), f.get("headline", ""),
            f.get("description", "") or loc.get("display", ""),
            f.get("status", ""),
            f.get("confidence_level", ""),
            str(f.get("first_seen_at", ""))[:19],
        ]
        for j, val in enumerate(row, 1):
            cell = ws.cell(i, j, val)
            cell.alignment = Alignment(wrap_text=True)
            cell.border = _border

    _apply_borders(ws)


def _sheet_exposure(wb, data: dict):
    ws = wb.create_sheet("Exposure")
    entities = data.get("entities", [])

    headers = ["Entity", "Type", "Risk", "Confidence", "First Seen", "Last Seen", "Sources"]
    _write_header(ws, "EXPOSED ASSET INVENTORY", headers)

    for i, e in enumerate(entities[:500], start=4):
        sources = e.get("sources", [])
        row = [
            e.get("entity_value", ""),
            e.get("entity_type", ""),
            e.get("risk_level", ""),
            e.get("confidence_level", ""),
            str(e.get("first_seen_at", ""))[:19],
            str(e.get("last_seen_at", ""))[:19],
            ", ".join(sources) if isinstance(sources, list) else str(sources),
        ]
        for j, val in enumerate(row, 1):
            cell = ws.cell(i, j, val)
            cell.alignment = Alignment(wrap_text=True)
            cell.border = _border

    _apply_borders(ws)


def _sheet_certificates(wb, data: dict):
    ws = wb.create_sheet("Certificates")
    certs = data.get("certificates", [])

    headers = ["Subject CN", "Issuer", "Risk", "Deployment", "Expires", "SANs"]
    _write_header(ws, "CERTIFICATE INVENTORY", headers)

    for i, c in enumerate(certs[:200], start=4):
        san_str = ", ".join((c.get("san_dns_names") or [])[:5])
        row = [
            c.get("subject_cn", ""),
            c.get("issuer_organization", ""),
            c.get("risk", ""),
            c.get("deployment_state", ""),
            str(c.get("not_after", ""))[:10],
            san_str,
        ]
        for j, val in enumerate(row, 1):
            cell = ws.cell(i, j, val)
            cell.alignment = Alignment(wrap_text=True)
            cell.border = _border

    _apply_borders(ws)


def _write_header(ws, title: str, headers: list[str]):
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(headers))
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color=INK)
    ws["A1"].fill = PatternFill("solid", fgColor=PRIMARY_DEEP)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for i, h in enumerate(headers, 1):
        cell = ws.cell(3, i, h)
        cell.font = Font(bold=True, color=PRIMARY)
        cell.fill = PatternFill("solid", fgColor=CARD_ALT)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _border


def _apply_borders(ws):
    for col in ws.columns:
        length = max(13, min(48, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = length
